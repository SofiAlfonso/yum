# Autor: Ana Sofía Alfonso


from typing import List
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO


from .report_interface import IReportGenerator


class ExcelReportGenerator(IReportGenerator):
    """
    Implementación concreta del generador de reportes en formato Excel.
    Utiliza openpyxl para crear hojas de cálculo profesionales.
    """
    
    def generate(self, recipes: List, filename: str) -> HttpResponse:
        """Genera un reporte Excel con las recetas proporcionadas."""
        response = HttpResponse(content_type=self.get_content_type())
        response['Content-Disposition'] = f'attachment; filename="{filename}.{self.get_file_extension()}"'
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Recetas"
        
        # Estilos
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='23b387', end_color='23b387', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        title_font = Font(name='Arial', size=16, bold=True, color='23b387')
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título del reporte
        ws.merge_cells('A1:F1')
        ws['A1'] = '🍽️ Reporte de Recetas - YUM'
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment
        
        # Fecha de generación
        ws.merge_cells('A2:F2')
        ws['A2'] = f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].font = Font(name='Arial', size=10, color='666666')
        
        # Resumen
        ws['A4'] = '📊 RESUMEN'
        ws['A4'].font = Font(name='Arial', size=12, bold=True, color='23b387')
        
        ws['A5'] = 'Total de Recetas:'
        ws['B5'] = len(recipes)
        ws['A6'] = 'Recetas con Reseñas:'
        ws['B6'] = sum(1 for r in recipes if r.review_count > 0)
        ws['A7'] = 'Promedio de Rating:'
        avg_rating = sum(r.avg_rating or 0 for r in recipes) / len(recipes) if recipes else 0
        ws['B7'] = f"{avg_rating:.1f} ⭐"
        
        # Estilo al resumen
        for row in range(5, 8):
            ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
            ws[f'B{row}'].font = Font(name='Arial', size=10)
        
        # Encabezados de la tabla (fila 9)
        headers = ['#', 'Título', 'Usuario', 'Categoría', 'Valor Nutricional', 'Porciones', 'Reseñas', 'Rating']
        ws.append([])  # Espacio
        header_row = 9
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style
        
        # Datos de las recetas
        for idx, recipe in enumerate(recipes, 1):
            row = [
                idx,
                recipe.title,
                recipe.user.username,
                recipe.get_category_display(),
                recipe.nutritional_value,
                recipe.portions,
                recipe.review_count,
                f"{recipe.avg_rating:.1f}" if recipe.avg_rating else "N/A"
            ]
            ws.append(row)
            
            # Estilo a la fila
            current_row = header_row + idx
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.alignment = Alignment(horizontal='center' if col_num in [1, 5, 6, 7, 8] else 'left')
                cell.border = border_style
                
                # Fondo alternado
                if idx % 2 == 0:
                    cell.fill = PatternFill(start_color='f5f5f5', end_color='f5f5f5', fill_type='solid')
        
        # Ajustar anchos de columna
        column_widths = {
            'A': 8,   # #
            'B': 35,  # Título
            'C': 20,  # Usuario
            'D': 20,  # Categoría
            'E': 18,  # Valor Nutricional
            'F': 12,  # Porciones
            'G': 12,  # Reseñas
            'H': 12,  # Rating
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Ajustar altura de filas
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[header_row].height = 20
        
        # Crear hoja de estadísticas adicionales
        stats_ws = wb.create_sheet("Estadísticas")
        stats_ws['A1'] = '📈 Estadísticas Detalladas'
        stats_ws['A1'].font = title_font
        stats_ws.merge_cells('A1:B1')
        
        # Categorías más populares
        from collections import Counter
        categories = Counter([r.get_category_display() for r in recipes])
        
        stats_ws['A3'] = 'Categorías Más Populares:'
        stats_ws['A3'].font = Font(name='Arial', size=12, bold=True)
        
        row = 4
        for category, count in categories.most_common():
            stats_ws[f'A{row}'] = category
            stats_ws[f'B{row}'] = count
            row += 1
        
        # Top usuarios con más recetas
        from collections import Counter
        users = Counter([r.user.username for r in recipes])
        
        stats_ws[f'A{row+1}'] = 'Top Usuarios con Más Recetas:'
        stats_ws[f'A{row+1}'].font = Font(name='Arial', size=12, bold=True)
        
        row += 2
        for username, count in users.most_common(5):
            stats_ws[f'A{row}'] = username
            stats_ws[f'B{row}'] = count
            row += 1
        
        # Ajustar anchos en hoja de estadísticas
        stats_ws.column_dimensions['A'].width = 30
        stats_ws.column_dimensions['B'].width = 15
        
        # Guardar el archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response.write(output.getvalue())
        output.close()
        return response
    
    def get_content_type(self) -> str:
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    def get_file_extension(self) -> str:
        return 'xlsx'